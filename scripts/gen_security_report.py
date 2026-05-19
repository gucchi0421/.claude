#!/usr/bin/env python3
"""
Claude Code セキュリティ設定 + 活用まとめ Excel 生成スクリプト
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── スタイル定数 ───────────────────────────────────────────

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(bold=bold, size=size, color=color)

def body_font(size=10, bold=False, color="333333"):
    return Font(bold=bold, size=size, color=color)

def wrap_align(horizontal="left", vertical="center"):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

COLORS = {
    "red_header":    "C0392B",
    "red_light":     "FADBD8",
    "orange_header": "D35400",
    "orange_light":  "FDEBD0",
    "green_header":  "1E8449",
    "green_light":   "D5F5E3",
    "blue_header":   "1A5276",
    "blue_light":    "D6EAF8",
    "purple_header": "6C3483",
    "purple_light":  "E8DAEF",
    "gray_header":   "424242",
    "gray_light":    "F5F5F5",
    "yellow_light":  "FEF9E7",
    "teal_header":   "0E6655",
    "teal_light":    "D1F2EB",
}


def apply_header_row(ws, row, cols, fg, bg):
    for col_idx, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.fill = fill(bg)
        c.font = header_font(color=fg)
        c.alignment = wrap_align("center")
        c.border = border()


def apply_data_row(ws, row, cols, bg="FFFFFF", bold_first=False):
    for col_idx, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.fill = fill(bg)
        c.font = Font(bold=(bold_first and col_idx == 1), size=10, color="333333")
        c.alignment = wrap_align()
        c.border = border()


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_section_title(ws, row, text, col_span, bg, fg="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, size=11, color=fg)
    c.alignment = wrap_align("left")
    c.border = border()


# ══════════════════════════════════════════════
#  Sheet 1: 全体概要
# ══════════════════════════════════════════════

def build_overview(wb):
    ws = wb.create_sheet("① 全体概要")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 20, 48, 24])

    # タイトル
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Claude Code 活用まとめ  ―  エージェント・スキル・セキュリティ設計の概要"
    t.fill = fill(COLORS["blue_header"])
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = wrap_align("center")
    t.border = border()
    ws.row_dimensions[1].height = 30

    # サブタイトル
    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = "本資料は Claude Code（Anthropic）をチーム業務に導入した際の構成・設計の概要を上長向けにまとめたものです。"
    s.fill = fill(COLORS["blue_light"])
    s.font = Font(size=10, color="1A5276")
    s.alignment = wrap_align("left")
    s.border = border()
    ws.row_dimensions[2].height = 22

    r = 4
    # ── エージェント一覧 ──
    add_section_title(ws, r, "■ 専門エージェント一覧（18種）", 4, COLORS["blue_header"])
    r += 1
    apply_header_row(ws, r, ["カテゴリ", "エージェント名", "主な役割", "自動連携先"], "FFFFFF", COLORS["gray_header"])
    ws.row_dimensions[r].height = 20
    r += 1

    agents = [
        ("マネジメント", "commander",      "全体統括。ユーザーのリクエストを分析し最適なエージェントへ自動振り分け・並列実行", "全エージェント"),
        ("マネジメント", "director",       "企画・要件定義・仕様書作成・施策優先順位付け",               "seo-analyst / coder"),
        ("マネジメント", "pm",             "タスク分解・工数見積・スケジュール作成・ボトルネック特定",     "director"),
        ("マネジメント", "recorder",       "セッション後の日報・作業サマリーを自動生成・保存",             "commander"),
        ("マネジメント", "sales-bridge",   "クライアントメールをディレクション用語に変換・回答文案作成",   "director / coder"),
        ("Web制作・開発", "coder",         "実装・バグ修正・コードレビュー・API設計",                      "security-analyst"),
        ("Web制作・開発", "designer",      "HTML/CSS実装・UI改善・デザインシステム整備",                   "coder"),
        ("Web制作・開発", "design-analyst","競合サイトのデザイントレンド調査・ベンチマーク分析",            "designer"),
        ("Web制作・開発", "security-analyst","XSS・SQLi・依存パッケージ脆弱性チェック・レポート生成",     "coder"),
        ("Web制作・開発", "wp-operator",   "WordPress管理画面操作（記事公開・プラグイン設定・バックアップ）","coder"),
        ("SEO・コンテンツ", "seo-analyst", "キーワード調査・競合分析・3C分析・SEO課題の優先度付け",        "writer / seo-engineer"),
        ("SEO・コンテンツ", "seo-engineer","メタタグ・構造化データ・AIOSEO設定・Core Web Vitals改善",      "wp-operator"),
        ("SEO・コンテンツ", "writer",      "SEO記事・LPコピー・リライト・FAQ執筆（品質パイプライン適用）",  "article-reviewer"),
        ("SEO・コンテンツ", "article-reviewer","記事品質スコアリング（80点以上で合格）・改善フィードバック","writer"),
        ("SEO・コンテンツ", "data-analyst","GA4・Search ConsoleデータのKPI分析・異常検知",                 "ad-manager / seo-analyst"),
        ("ビジネス", "accountant",         "請求書・振替一覧の金額照合・未入金検出・見積概算作成",          "—"),
        ("ビジネス", "ad-manager",         "Google/Meta広告レポート分析・改善提案・予算配分",              "data-analyst"),
        ("ビジネス", "notifier",           "Chatworkへの作業完了通知・承認依頼の自動送信",                 "commander"),
    ]

    cat_colors = {
        "マネジメント":   COLORS["blue_light"],
        "Web制作・開発":  COLORS["green_light"],
        "SEO・コンテンツ": COLORS["orange_light"],
        "ビジネス":       COLORS["purple_light"],
    }
    for ag in agents:
        bg = cat_colors.get(ag[0], "FFFFFF")
        apply_data_row(ws, r, ag, bg)
        ws.row_dimensions[r].height = 28
        r += 1

    # ── スキル一覧 ──
    r += 1
    add_section_title(ws, r, "■ スキル（スラッシュコマンド）一覧", 4, COLORS["teal_header"])
    r += 1
    apply_header_row(ws, r, ["コマンド", "トリガー例", "内容", "外部連携"], "FFFFFF", COLORS["gray_header"])
    ws.row_dimensions[r].height = 20
    r += 1

    skills = [
        ("/chatwork",         "/chatwork rid311735252 <<EOF…",  "指定Chatworkルームにメッセージ送信。.envのAPIキーを自動参照",               "Chatwork API"),
        ("/ftp-deploy",       "/ftp-deploy",                    "変更ファイルをFTPで本番へ自動アップ。バックアップ→pending確認→実行の3ステップ", "FTP サーバー"),
        ("/ga4-fetch",        "/ga4-fetch",                     "GA4 Data APIからデータ取得→JSON出力→data-analystへ渡す",                    "Google Analytics"),
        ("/seo-log",          "/seo-log",                       "SEO作業履歴をGoogleスプレッドシートへ自動記録",                              "Google Sheets"),
        ("/ut-ai-news-report","/ut-ai-news-report（スケジュール）","毎平日12時に4サイトを並列スクレイプ→カテゴリ分類→Chatwork自動投稿",     "Chatwork"),
        ("/wp-seo-manager",   "/wp-seo-manager",                "WordPressサイトのSEO対策を一気通貫で自動化（調査→承認→実行→Notion記録）",   "WP管理画面 / Notion"),
        ("/wp-manual-create", "/wp-manual-create",              "WP管理画面をクロールしてスクショ付きPPTXマニュアルを自動生成",               "WP管理画面 / PowerPoint"),
    ]
    for sk in skills:
        apply_data_row(ws, r, sk, COLORS["teal_light"])
        ws.row_dimensions[r].height = 28
        r += 1


# ══════════════════════════════════════════════
#  Sheet 2: セキュリティ設定（操作制御）
# ══════════════════════════════════════════════

def build_security(wb):
    ws = wb.create_sheet("② セキュリティ設定")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [18, 36, 40, 20])

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Claude Code セキュリティ設定  ―  操作制御・禁止コマンド・ファイル保護"
    t.fill = fill(COLORS["red_header"])
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = wrap_align("center")
    t.border = border()
    ws.row_dimensions[1].height = 30

    r = 3

    # ── DENY（完全禁止） ──
    add_section_title(ws, r, "■ 完全禁止（deny）― settings.json に定義", 4, COLORS["red_header"])
    r += 1
    apply_header_row(ws, r, ["種別", "パターン", "禁止理由", "回避策"], "FFFFFF", "B03A2E")
    r += 1

    deny_rows = [
        ("ファイル読み取り禁止", "Read(**/.env)\nRead(**/.env.*)\nRead(**/*.env)",
         "FTPパスワード・APIキー等の機密情報が含まれる環境変数ファイルをClaudeが参照できないよう完全ブロック",
         "接続情報は.envに分離。スクリプト側のみos.environ経由で参照"),
        ("ファイル読み取り禁止", "Read(**/*.pem)\nRead(**/*secret*)",
         "SSL証明書・秘密鍵・シークレットファイルの読み取りをブロック",
         "証明書は別管理。Claudeには渡さない"),
        ("ファイル編集禁止",   "Edit(**/.env)\nEdit(**/.env.*)",
         ".envを直接書き換えることによる意図しない認証情報の変更を防止",
         "環境変数の変更は人手で実施"),
        ("CI/CD編集禁止",     "Edit(/.github/workflows/**)",
         "GitHub ActionsワークフローをClaudeが改ざんしてサプライチェーン攻撃の踏み台になるリスクを防止",
         "CI設定の変更は開発者が直接編集"),
        ("破壊的Gitコマンド禁止", "Bash(git push -f *)\nBash(git push --force *)\nBash(git reset --hard *)",
         "強制プッシュ・履歴破壊によるリモートリポジトリの損壊を防止",
         "通常のgit pushのみ許可。reset --hardはユーザー確認後のみ"),
        ("ファイル一括削除禁止", "Bash(rm -rf *)",
         "再帰削除による誤ったディレクトリ全消しを防止",
         "ファイル単体のrmはフックで5秒遅延＋確認を挟む"),
        ("ネットワーク直接禁止", "Bash(curl *)\nBash(wget *)",
         "Claudeが任意URLへデータを送信・外部スクリプトをダウンロードするリスクを遮断",
         "外部通信はWebFetchツール（ホワイトリスト制）経由に限定"),
    ]
    for row in deny_rows:
        apply_data_row(ws, r, row, COLORS["red_light"])
        ws.row_dimensions[r].height = 42
        r += 1

    r += 1
    # ── ASK（要確認） ──
    add_section_title(ws, r, "■ 実行前にユーザー確認が必要（ask）", 4, COLORS["orange_header"])
    r += 1
    apply_header_row(ws, r, ["種別", "パターン", "確認が必要な理由", "備考"], "FFFFFF", "A04000")
    r += 1

    ask_rows = [
        ("ファイル書き込み",    "Write",                   "新規ファイル作成は意図しない上書きを防ぐため承認必須",             "—"),
        ("ファイル編集",        "Edit",                    "既存コードの変更は必ずdiff確認後に実施",                           "—"),
        ("コミット",           "Bash(git commit*)",        "コミットメッセージ・対象ファイルを人が確認してから実行",            "--no-verify禁止"),
        ("プッシュ",           "Bash(git push*)",          "リモートへの反映は必ず人が承認",                                   "force push禁止"),
        ("npm公開",            "Bash(npm publish*)",       "パッケージ公開は取り消し不能なため確認必須",                       "—"),
        ("本番デプロイ",       "Bash(wrangler pages deploy*)", "本番環境へのデプロイは承認後のみ実行",                         "—"),
    ]
    for row in ask_rows:
        apply_data_row(ws, r, row, COLORS["orange_light"])
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    # ── ALLOW（安全な操作は自動許可） ──
    add_section_title(ws, r, "■ 自動許可（allow）― 安全な読み取り・確認系のみ", 4, COLORS["green_header"])
    r += 1
    apply_header_row(ws, r, ["許可パターン", "内容", "許可理由", "—"], "FFFFFF", "1A6630")
    r += 1

    allow_rows = [
        ("Read / Glob / Grep",          "ファイル読み取り・検索（.env等除く）",           "コード理解・調査のみで変更なし"),
        ("WebFetch / WebSearch",         "Web参照・検索",                                 "外部通信はツール制御内で完結"),
        ("Bash(ls / find / stat / wc)", "ディレクトリ確認・ファイル情報取得",             "破壊的操作なし"),
        ("Bash(git status/diff/log)",   "Git状態確認（読み取りのみ）",                   "変更を伴わない参照系のみ"),
        ("Bash(docker / docker-compose)", "コンテナ操作（開発環境用）",                  "本番サーバー直接操作ではない"),
        ("Bash(npm run *)",             "スクリプト実行（ビルド・テスト等）",             "package.jsonで管理された安全なスクリプト"),
        ("Bash(* --version / --help)",  "バージョン確認・ヘルプ表示",                    "情報取得のみ"),
    ]
    for row in allow_rows:
        c1, c2, c3 = row
        apply_data_row(ws, r, (c1, c2, c3, ""), COLORS["green_light"])
        ws.row_dimensions[r].height = 26
        r += 1


# ══════════════════════════════════════════════
#  Sheet 3: フック（自動ガード）
# ══════════════════════════════════════════════

def build_hooks(wb):
    ws = wb.create_sheet("③ 自動ガード（フック）")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [20, 30, 50, 18])

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Claude Code フック設定  ―  ツール実行前後に自動で割り込む安全装置"
    t.fill = fill(COLORS["orange_header"])
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = wrap_align("center")
    t.border = border()
    ws.row_dimensions[1].height = 30

    r = 3
    # PreToolUse
    add_section_title(ws, r, "■ PreToolUse（ツール実行直前に割り込む）", 4, COLORS["orange_header"])
    r += 1
    apply_header_row(ws, r, ["トリガー（matcher）", "実行されるコマンド", "何を防いでいるか", "失敗時の挙動"], "FFFFFF", "A04000")
    r += 1

    pre_rows = [
        ("Bash(git add*)",
         "git diff --cached --name-only | grep -E '(^|/)\\.env'\n→ ヒットしたら exit 1",
         ".envファイルが誤ってステージングされていないかチェック。\nコミット前の最後の砦として機密情報の流出を防ぐ",
         "exit 1でgit addを強制中断。「.envをステージしようとしています」と警告表示"),
        ("Bash(rm*)",
         "echo '⚠️ 削除コマンド検出' && sleep 5",
         "ファイル削除を5秒間遅延させ、誤操作の取り消し猶予を確保する\n（rm -rfは別途denyで完全禁止）",
         "Ctrl+Cで中断可能。5秒後に実行継続"),
    ]
    for row in pre_rows:
        apply_data_row(ws, r, row, COLORS["orange_light"])
        ws.row_dimensions[r].height = 55
        r += 1

    r += 1
    # PostToolUse
    add_section_title(ws, r, "■ PostToolUse（ツール実行直後に自動で動く）", 4, COLORS["blue_header"])
    r += 1
    apply_header_row(ws, r, ["トリガー（matcher）", "実行されるスクリプト", "何をしているか", "出力先"], "FFFFFF", "1A5276")
    r += 1

    post_rows = [
        ("Edit | Write | Bash(rm *)",
         "python3 ~/.claude/scripts/log_work.py",
         "ファイル編集・作成・削除をリアルタイムで自動記録。\n①セッションログ（作業日誌）への追記\n②FTPデプロイ用のpendingリストの自動更新",
         ".claude/logs/work/\n.claude/logs/deploy/.pendings.md"),
        ("Bash(claude plugin install*)",
         "python3 ~/.claude/scripts/add_plugin_to_setup.py",
         "プラグインをインストールした際、設定ファイルへの追記を自動化。\n手作業の記録漏れを防ぐ",
         "settings.json"),
        ("Bash(claude mcp add*)",
         "python3 ~/.claude/scripts/add_mcp_to_setup.py",
         "MCPサーバー追加時に設定ファイルへ自動反映。\n複数環境での設定同期のため",
         "settings.json"),
    ]
    for row in post_rows:
        apply_data_row(ws, r, row, COLORS["blue_light"])
        ws.row_dimensions[r].height = 55
        r += 1

    r += 1
    # 自動作業ログの説明
    add_section_title(ws, r, "■ 自動作業ログの仕組み（log_work.py）", 4, COLORS["gray_header"])
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+5, end_column=4)
    c = ws.cell(row=r, column=1)
    c.value = (
        "Edit・Write・rmが実行されるたびに以下の2つのファイルを自動更新する。\n\n"
        "【セッションログ】 .claude/logs/work/YYYYMMDDHHMMSS_session.md\n"
        "  → 時刻・操作種別・ファイル名・diff（前後3行のコンテキスト付き）を記録\n"
        "  → 1日1ファイル。同日の複数セッションは同一ファイルに追記\n\n"
        "【デプロイpendingリスト】 .claude/logs/deploy/.pendings.md\n"
        "  → 同じファイルへの複数回操作は最新の操作（Edit/Create/Delete）で上書き\n"
        "  → /ftp-deploy 実行時にこのリストを読み込んで本番へアップロード\n"
        "  → デプロイ完了後にリセット"
    )
    c.fill = fill(COLORS["gray_light"])
    c.font = Font(size=10, color="333333")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = border()
    ws.row_dimensions[r].height = 130


# ══════════════════════════════════════════════
#  Sheet 4: 機密情報管理（FTP/SSH）
# ══════════════════════════════════════════════

def build_credentials(wb):
    ws = wb.create_sheet("④ 機密情報管理（FTP・SSH）")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 28, 50, 18])

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "機密情報の管理設計  ―  パスワード・APIキーをClaudeに見せずに外部接続する方法"
    t.fill = fill(COLORS["purple_header"])
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = wrap_align("center")
    t.border = border()
    ws.row_dimensions[1].height = 30

    r = 3

    # ── 設計思想 ──
    add_section_title(ws, r, "■ 基本設計思想", 4, COLORS["purple_header"])
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=4)
    c = ws.cell(row=r, column=1)
    c.value = (
        "Claude（AI）は接続先のパスワード・APIキーを「一切知らない」設計になっている。\n\n"
        "【問題】 AIが機密情報を知ると → プロンプトに含まれる → ログ・会話履歴に残るリスクがある\n"
        "【解決】 接続情報は .env ファイルに分離し、Pythonスクリプトが実行時にのみ参照する\n"
        "         Claude は .env の「存在確認（ls）」だけ行い、中身は読まない（denyで強制）"
    )
    c.fill = fill(COLORS["purple_light"])
    c.font = Font(size=10, color="333333")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = border()
    ws.row_dimensions[r].height = 85
    r += 4

    r += 1
    # ── FTP接続の仕組み ──
    add_section_title(ws, r, "■ FTP接続（ftp_deploy.py）の機密情報管理", 4, COLORS["blue_header"])
    r += 1
    apply_header_row(ws, r, ["ステップ", "誰が・何を", "機密情報との関係", "設定ファイル"], "FFFFFF", "1A5276")
    r += 1

    ftp_rows = [
        ("① .envに接続情報を記載",
         "開発者が手動で作成",
         "CLAUDE_FTP_HOST / CLAUDE_FTP_USERNAME / CLAUDE_FTP_PASSWORD を記載\n★ Claudeはこのファイルを絶対に読まない（denyルール）",
         "プロジェクト直下/.env"),
        ("② パスマッピング設定",
         "Claude が初回セットアップ時に作成",
         "ローカルパス→リモートパスの対応表のみ。パスワードは含まない",
         ".claude/settings.ftp.json"),
        ("③ pendingリストを確認",
         "Claude が .pendings.md を表示してユーザーに確認",
         "アップロード対象ファイルの一覧のみ。認証情報なし",
         ".claude/logs/deploy/.pendings.md"),
        ("④ ユーザーが承認",
         "人間が「実行してください」と指示",
         "この段階でもClaudeは認証情報を知らない",
         "—"),
        ("⑤ Pythonスクリプトが実行",
         "ftp_deploy.py が os.environ 経由で .env を参照",
         ".envはスクリプト内でのみ読み込まれ、Claudeの会話には戻ってこない\nFTPS（暗号化）で接続",
         "~/.claude/scripts/ftp_deploy.py"),
        ("⑥ バックアップ→アップロード",
         "スクリプトが自動でバックアップ後にアップ",
         "認証情報は接続時のみ使用。完了後メモリから破棄",
         "リモート/.claude/backup/"),
    ]
    for row in ftp_rows:
        apply_data_row(ws, r, row, COLORS["blue_light"])
        ws.row_dimensions[r].height = 48
        r += 1

    r += 1
    # ── SSH設定 ──
    add_section_title(ws, r, "■ SSH接続設定の管理方法", 4, COLORS["teal_header"])
    r += 1
    apply_header_row(ws, r, ["設定項目", "値（例）", "説明", "機密情報の扱い"], "FFFFFF", "0E6655")
    r += 1

    ssh_rows = [
        ("connection（接続名）", "mysite",
         "~/.ssh/config に定義した Host 名を指定。\nパスワードは不要（公開鍵認証）",
         "パスワード不使用"),
        ("config（設定ファイル）", "~/.ssh/config",
         "標準のSSH設定ファイルを参照。\nClaudeは設定ファイルのパスを知るだけで中身は読まない",
         "Claudeは中身を読まない"),
        ("allowed_dirs（操作可能ディレクトリ）", "/var/www/html",
         "Claudeがサーバー上で操作できるディレクトリをホワイトリストで制限。\n設定外のパスへのアクセスをブロック",
         "スコープを最小化"),
    ]
    for row in ssh_rows:
        apply_data_row(ws, r, row, COLORS["teal_light"])
        ws.row_dimensions[r].height = 44
        r += 1

    r += 1
    # ── Chatwork / GA4 / その他APIキー ──
    add_section_title(ws, r, "■ その他APIキー・認証情報の管理方法", 4, COLORS["gray_header"])
    r += 1
    apply_header_row(ws, r, ["サービス", "キーの格納場所", "Claudeの関与", "スクリプト参照方法"], "FFFFFF", "424242")
    r += 1

    api_rows = [
        ("Chatwork API",      "プロジェクト直下/.env\n（CLAUDE_CHATWORK_API_KEY）",
         "存在確認（ls）のみ。中身は読まない",
         "chatwork_send.py が os.environ 経由で参照"),
        ("Google Analytics (GA4)", "~/.claude/.env\n（CLAUDE_GCP_CREDENTIALS=/path/to/key.json）",
         "鍵ファイルのパスのみ知る。ファイル内容は読まない",
         "ga4_fetch.py がサービスアカウントJSONを直接読み込み"),
        ("Google Sheets (SEOログ)", "~/.claude/.env\n（CLAUDE_GCP_CREDENTIALS 共用）",
         "同上",
         "seo_log.py がGCP認証情報を使用"),
        ("WordPress ログイン情報", "~/wp-credentials/[案件名].json",
         "ブラウザ操作（Playwright）でログインフォームに入力。\nClaudeの会話には戻ってこない",
         "wp-manual-createスキルがファイルを読み込みブラウザ経由で使用"),
        ("FTP接続情報",       "プロジェクト直下/.env\n（CLAUDE_FTP_HOST 等）",
         "存在確認のみ。中身は読まない（denyで強制）",
         "ftp_deploy.py が実行時にのみ参照"),
    ]
    for row in api_rows:
        apply_data_row(ws, r, row, COLORS["gray_light"])
        ws.row_dimensions[r].height = 44
        r += 1


# ══════════════════════════════════════════════
#  Sheet 5: コード品質・セキュリティルール
# ══════════════════════════════════════════════

def build_code_rules(wb):
    ws = wb.create_sheet("⑤ コード品質ルール")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [20, 38, 42, 18])

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "コード品質・セキュリティルール  ―  全プロジェクト共通で適用されるコーディング規約"
    t.fill = fill(COLORS["green_header"])
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = wrap_align("center")
    t.border = border()
    ws.row_dimensions[1].height = 30

    r = 3
    add_section_title(ws, r, "■ セキュリティコーディングルール（OWASP Top 10 対応）", 4, COLORS["red_header"])
    r += 1
    apply_header_row(ws, r, ["脆弱性種別", "禁止事項", "必須対応", "適用言語"], "FFFFFF", "B03A2E")
    r += 1

    sec_rows = [
        ("XSS（クロスサイトスクリプティング）",
         "ユーザー入力をそのままHTMLに出力",
         "必ずサニタイズ・エスケープしてから出力",
         "PHP / JS / TypeScript"),
        ("SQLインジェクション",
         "文字列結合でSQLクエリを組み立てる",
         "プリペアドステートメント必須",
         "PHP / Go"),
        ("認証不備",
         "認証チェックをバイパスする実装",
         "全リクエストで認証状態を検証",
         "全言語"),
        ("シークレットのハードコード",
         "APIキー・パスワードをコードに直接記載",
         "環境変数 / wp-config.php で管理",
         "全言語"),
        ("CSRF（クロスサイトリクエストフォージェリ）",
         "nonceなしのフォーム・Ajaxリクエスト",
         "フォーム・Ajaxには必ずnonce付与と検証",
         "PHP（WordPress）/ JS"),
        ("オープンリダイレクト",
         "ユーザー入力URLへ無制限にリダイレクト",
         "リダイレクト先URLはホワイトリストで制限",
         "全言語"),
    ]
    for row in sec_rows:
        apply_data_row(ws, r, row, COLORS["red_light"])
        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    add_section_title(ws, r, "■ Git ワークフロールール（Conventional Commits）", 4, COLORS["blue_header"])
    r += 1
    apply_header_row(ws, r, ["ルール", "内容", "理由", "—"], "FFFFFF", "1A5276")
    r += 1

    git_rows = [
        ("コミット形式",
         "<type>(<scope>): <summary>\n\n例: feat(contact): add form validation",
         "Conventional Commits準拠でCHANGELOG自動生成・レビューしやすさを確保"),
        ("summaryは英語",
         "動詞原形で始める（add / fix / update 等）",
         "グローバルなコントリビューターが読める・検索しやすい"),
        ("bodyは日本語",
         "「何をしたか」でなく「なぜしたか」を記述",
         "後から読む人がコンテキストを理解できるようにする"),
        ("自動コミット禁止",
         "Claudeは確認なしにコミット・プッシュしない",
         "意図しない変更の混入を防ぐ"),
        ("ブランチ命名規則",
         "feat/ fix/ chore/ refactor/ + kebab-case\n例: feat/add-contact-form",
         "PRレビューで変更の種類が一目でわかる"),
    ]
    for row in git_rows:
        c1, c2, c3 = row
        apply_data_row(ws, r, (c1, c2, c3, ""), COLORS["blue_light"])
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1
    add_section_title(ws, r, "■ コーディング規約（言語別）", 4, COLORS["teal_header"])
    r += 1
    apply_header_row(ws, r, ["言語", "管理ファイル", "主な規約内容", "—"], "FFFFFF", "0E6655")
    r += 1

    lang_rows = [
        ("PHP",        "rules/code-styles/php.md",        "WordPress準拠・プリペアドステートメント必須・nonce検証"),
        ("TypeScript", "rules/code-styles/typescript.md", "型安全必須・any禁止・関数型スタイル推奨"),
        ("JavaScript", "rules/code-styles/javascript.md", "ESModules・async/await推奨・eval禁止"),
        ("Go",         "rules/code-styles/go.md",         "gofmt準拠・エラーは必ずハンドル・goroutine管理"),
        ("HTML",       "rules/code-styles/html.md",       "セマンティックHTML・alt必須・aria属性"),
        ("CSS/SCSS",   "rules/code-styles/css-scss.md",   "BEM命名・変数で色管理・!important禁止"),
        ("WordPress",  "rules/code-styles/wordpress.md",  "WordPress Coding Standards準拠・フック経由で拡張"),
    ]
    for row in lang_rows:
        c1, c2, c3 = row
        apply_data_row(ws, r, (c1, c2, c3, ""), COLORS["teal_light"])
        ws.row_dimensions[r].height = 26
        r += 1


# ══════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════

def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    build_overview(wb)
    build_security(wb)
    build_hooks(wb)
    build_credentials(wb)
    build_code_rules(wb)

    output = Path.home() / "documents" / "Claude活用まとめ.xlsx"
    wb.save(output)
    print(f"✅ 保存完了: {output}")


if __name__ == "__main__":
    main()
