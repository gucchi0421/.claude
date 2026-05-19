"""
SEO対策計画レポートをExcelで生成するスクリプト。
Claude が audit/research/plan の結果を引数で渡して呼び出す。

使い方:
  python generate_report.py --data '{"audit": {...}, "plan": [...]}'
  または Claude が直接 data dict を渡してインポートして使う。
"""

import sys
import json
import argparse
import tempfile
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が未インストールです。pip install openpyxl を実行してください。")
    sys.exit(1)


# スクリプトの場所に関係なくOSの一時ディレクトリに出力する
OUTPUT_DIR = Path(tempfile.gettempdir())


# --- スタイル定義 ---

def header_style(ws, row, col, value, bg="1F4E79", fg="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell

def data_cell(ws, row, col, value, bold=False, wrap=True, color=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or "000000")
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell

def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


# --- シート1: 調査結果 ---

def build_audit_sheet(ws, audit: dict):
    ws.title = "調査結果"
    ws.row_dimensions[1].height = 24

    headers = ["項目", "内容", "評価", "備考"]
    for i, h in enumerate(headers, 1):
        header_style(ws, 1, i, h)

    rows = audit.get("rows", [])
    if not rows:
        rows = [
            ("GSC平均順位", audit.get("avg_rank", "-"), "-", ""),
            ("対策KW順位", audit.get("kw_rank", "-"), "-", ""),
            ("AIOSEOスコア（最低）", audit.get("min_score", "-"), "-", ""),
            ("PageSpeed（モバイル）", audit.get("pagespeed_mobile", "-"), "-", ""),
            ("PageSpeed（デスクトップ）", audit.get("pagespeed_desktop", "-"), "-", ""),
            ("404エラー件数", audit.get("errors_404", "-"), "-", ""),
            ("孤立ページ数", audit.get("orphan_pages", "-"), "-", ""),
            ("内部リンクなし記事数", audit.get("no_internal_link", "-"), "-", ""),
        ]

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            data_cell(ws, r_idx, c_idx, val)

    auto_width(ws)


# --- シート2: 作業計画 ---

def build_plan_sheet(ws, plan: list):
    ws.title = "作業計画"
    ws.row_dimensions[1].height = 24

    headers = ["優先度", "対象ページ/設定", "変更種別", "変更内容サマリー", "管理場所"]
    for i, h in enumerate(headers, 1):
        header_style(ws, 1, i, h)

    priority_colors = {"A": "FF0000", "B": "FF9900", "C": "007B00"}

    for r_idx, item in enumerate(plan, 2):
        priority = item.get("priority", "-")
        color = priority_colors.get(priority, "000000")
        data_cell(ws, r_idx, 1, f"優先{priority}", bold=True, color=color)
        data_cell(ws, r_idx, 2, item.get("target", ""))
        data_cell(ws, r_idx, 3, item.get("type", ""))
        data_cell(ws, r_idx, 4, item.get("summary", ""))
        data_cell(ws, r_idx, 5, item.get("location", ""))

    auto_width(ws)


# --- シート3: 変更Diff ---

def build_diff_sheet(ws, diffs: list):
    ws.title = "変更Diff"
    ws.row_dimensions[1].height = 24

    headers = ["対象ページ", "変更箇所", "変更前", "変更後", "管理場所"]
    for i, h in enumerate(headers, 1):
        header_style(ws, 1, i, h)

    red_fill = PatternFill(fill_type="solid", fgColor="FFE0E0")
    green_fill = PatternFill(fill_type="solid", fgColor="E0FFE0")

    for r_idx, diff in enumerate(diffs, 2):
        data_cell(ws, r_idx, 1, diff.get("page", ""))
        data_cell(ws, r_idx, 2, diff.get("field", ""))

        before_cell = data_cell(ws, r_idx, 3, diff.get("before", ""))
        before_cell.fill = red_fill

        after_cell = data_cell(ws, r_idx, 4, diff.get("after", ""))
        after_cell.fill = green_fill

        data_cell(ws, r_idx, 5, diff.get("location", ""))

    auto_width(ws)


# --- メイン ---

def generate(data: dict) -> Path:
    today = date.today().strftime("%Y%m%d")
    output_path = OUTPUT_DIR / f"seo_report_{today}.xlsx"

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    build_audit_sheet(wb.create_sheet(), data.get("audit", {}))
    build_plan_sheet(wb.create_sheet(), data.get("plan", []))
    build_diff_sheet(wb.create_sheet(), data.get("diffs", []))

    wb.save(output_path)
    print(f"生成完了: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", type=str, default="{}")
    args = parser.parse_args()
    data = json.loads(args.data)
    generate(data)
